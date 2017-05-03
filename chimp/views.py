
from django.views import generic
from django.contrib.auth import authenticate,login,logout
from django import forms
from django.shortcuts import render , redirect
from . forms import LoginForm, RegistrationForm, EmailForm, AddCampaignForm, EditCampaignForm, EditUserProfileForm
from django.core.urlresolvers import reverse_lazy
from django.http import HttpResponseRedirect, HttpResponse
from django.contrib.auth.models import User
from . models import MailingList, TemplateList, Campaign, UserProfile
from django.contrib.messages.views import SuccessMessageMixin
from django.contrib import messages
from django.views.generic.edit import CreateView, UpdateView, DeleteView
from django.views import View
from openpyxl import load_workbook
from django.core.mail import send_mail
from django.template.loader import render_to_string
from django.core.validators import validate_email
from django.core.exceptions import ValidationError
from django.utils import timezone



class LoginView(generic.FormView):			# User login view
	template_name='chimp/login.html'
	form_class=LoginForm
	success_url=reverse_lazy('chimp:dashboard')

	def get(self, *args, **kwargs):
		if self.request.user:
			if self.request.user.is_authenticated():
				redirect('chimp:dashboard')
		return super(LoginView, self).get(*args, **kwargs)

	def form_valid(self,form):
		username =form.cleaned_data.get('username')
		password=form.cleaned_data.get('password')
		user=authenticate(username=username,password=password)
		if user.is_active:
			login(self.request, user)
			return redirect(self.success_url)


class LogoutView(generic.View):

	def get(self,request):
		logout(request)
		messages.success(request, 'logout successfully')
		return redirect('chimp:login')


class RegistrationView(generic.FormView):
	template_name='chimp/registration.html'
	form_class=RegistrationForm
	success_url=reverse_lazy('chimp:login')

	def get(self, *args, **kwargs):
		if self.request.user.is_authenticated():
			return redirect('chimp:dashboard')
		return super(RegistrationView, self).get(*args, **kwargs)

	def form_valid(self,form):
		username=form.cleaned_data.get('username')
		name =form.cleaned_data.get('name')
		password=form.cleaned_data.get('password')
		business_name =form.cleaned_data.get('business_name')
		email=form.cleaned_data.get('email')
		package=form.cleaned_data.get('package')
		first_name=name.split()[0]
		last_name = ''

		try:
			last_name=name.split()[1]
		except:
			pass

		user=User(username=username, first_name=first_name, last_name=last_name, email=email)
		if User.objects.filter(username=user.username).exists():
			"""
			check whether user is not already exist
			"""
			messages.error(self.request,'Username already exist')
			return redirect('chimp:registration')

		user.set_password(password)
		user.save()
		userprofile=UserProfile(user=user,business_name=business_name,package=package , email=email)
		userprofile.save()
		messages.success(self.request, 'User Register Successfully')
		return redirect(self.success_url)




class AddCampaignView(generic.FormView):
	template_name='chimp/create.html'
	success_url=reverse_lazy('chimp:dashboard')
	form_class = AddCampaignForm

	def get_form_kwargs(self):
		kwargs = super(AddCampaignView, self).get_form_kwargs()
		kwargs['request'] = self.request
		return kwargs

	def form_valid(self, form, *args, **kwargs):
		if form.is_valid():
			campaign = form.save(commit=False)
			campaign.user = self.request.user
			form.save()
			messages.success(self.request, 'Campaign created  Successfully')
			return super(AddCampaignView, self).form_valid(form, *args, **kwargs)


class EditCampaignView(SuccessMessageMixin, generic.UpdateView):
	model=Campaign
	form_class = EditCampaignForm
	template_name='chimp/edit_campaign.html'
	success_url = reverse_lazy('chimp:show_campaign')
	success_message = 'Campaign Edited Successfully'

	def get_form_kwargs(self):
		kwargs = super(EditCampaignView, self).get_form_kwargs()
		kwargs['request'] = self.request
		return kwargs



class AddMailList(SuccessMessageMixin, CreateView):
	model= MailingList
	fields = ['name', 'mailing_list_path']
	#fields = '__all__'
	template_name = 'chimp/add_mail_list.html'
	success_message = 'Mail list added.'
	success_url=reverse_lazy('chimp:show_mailing_list')
	

	def form_valid(self,form, *args, **kwargs):
		if form.is_valid():
			mailinglist=form.save(commit=False)
			mailinglist.user=self.request.user
			file_type = self.request.FILES['mailing_list_path'].content_type
			mime_type_list=['application/vnd.ms-excel','application/msexcel','application/x-msexcel','application/x-ms-excel','application/x-excel','application/x-dos_ms_excel','application/xls','application/x-xls','application/vnd.openxmlformats-officedocument.spreadsheetml.sheet','application/vnd.ms-excel.sheet.binary.macroEnabled.12','application/vnd.openxmlformats-officedocument.spreadsheetml.template',]

			try:
				if(file_type in mime_type_list):
					mailinglist.save()
				else:
					return HttpResponse("wrong file format")
			except:
				raise forms.ValidationError('Wrong file format')

			return super(AddMailList, self).form_valid(form, *args, **kwargs)


class AddTemplateList(SuccessMessageMixin, CreateView):
	model=TemplateList
	fields= ['image','template_name','category','path']
	template_name='chimp/add_template_list.html'
	success_message='Template list added'
	success_url='/admin'


class SendEmailView(View):

	def get(self, request, *args, **kwargs):
		campaign_obj=Campaign.objects.get(id=self.kwargs['pk'])
		u=UserProfile.objects.get(user=self.request.user)
		print("+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++")
		print(campaign_obj.mailing_list.mailing_list_path)
		path=campaign_obj.mailing_list.mailing_list_path
		template_path=campaign_obj.template.file
		campaign_name=campaign_obj.name

		wb = load_workbook(filename = path)
		sheet = wb.worksheets[0]
		row_count = sheet.max_row
		column_count = sheet.max_column
		users_name=[]
		receivers=[]
		for i in range(2, row_count):
			for j in range(1, column_count):
				email_o = sheet.cell(row=i, column=2).value
				try:
					validate_email(email_o)
					receivers.append(email_o)
					users_name.append(sheet.cell(row=i, column=1).value)
				except ValidationError as e:
					print(sheet.cell(row=i, column=1).value + " is invalid email id\n")

		sent_mail=len(users_name)	# Number of count of mailing list

		rem_mail=u.remaining_email
		if( sent_mail < rem_mail):
			msg_html = render_to_string(template_path, {'content': self.request.user})
			for receiver in receivers:
				send_mail(campaign_name, '', u.email, [receiver], html_message=msg_html,)
			total_sent_email=u.sent_email+sent_mail
			total_remaining_email=u.remaining_email-sent_mail
			u.sent_email=total_sent_email
			u.remaining_email=total_remaining_email
			u.save()
			return HttpResponse("Email sent succesfuly")
		else:
			return HttpResponse("Emails are less")

class DashboardView(generic.ListView):
	template_name='chimp/dashboard.html'
	model=Campaign

	def get(self, *args, **kwargs):
		#import ipdb; ipdb.set_trace()
		if not self.request.user.is_authenticated():
			return redirect('chimp:login')
		return super(DashboardView, self).get(*args, **kwargs)

	def get_queryset(self):
		return Campaign.objects.filter(user=self.request.user)

class HomepageView(generic.ListView):
	template_name='chimp/homepage.html'

	def get_queryset(self):
		return TemplateList.objects.all()

class CampaignDetailView(generic.DetailView):
	template_name = 'chimp/campaign_detail.html'
	model = Campaign

	def get_object(self):
		# Call the superclass
		object = super(CampaignDetailView, self).get_object()
		# Record the last accessed date
		object.last_accessed = timezone.now()
		object.save()
		# Return the object
		return object

class ShowTemplateView(generic.ListView):
	template_name='chimp/show_template.html'
	model=TemplateList

	def get_queryset(self):
		return TemplateList.objects.all()

class ShowMailingListView(generic.ListView):
	template_name='chimp/show_mailing_list.html'

	def get_queryset(self):
		return MailingList.objects.filter(user=self.request.user)

class ShowCampaignView(generic.ListView):
	template_name='chimp/show_campaign.html'

	def get_queryset(self):
		return Campaign.objects.filter(user=self.request.user)


class UserProfileView(generic.ListView):
	template_name ='chimp/user_profile.html'

	def get_queryset(self):
		return UserProfile.objects.filter(user=self.request.user)


class DeleteCampaignView(generic.DeleteView):
	model=Campaign
	template_name='chimp/delete_campaign.html'
	success_url = reverse_lazy('chimp:show_campaign')

	def delete(self, request, *args, **kwargs):
		"""
		Calls the delete() method on the fetched object and then
		redirects to the success URL.
		"""
		self.object = self.get_object()
		success_url = self.get_success_url()
		self.object.delete()
		messages.success(request, 'Campaign Deleted Successfully')
		return HttpResponseRedirect(success_url)


class DeleteMailingListView(generic.DeleteView):
	model=Campaign
	template_name='chimp/delete_mailing_list.html'
	success_url = reverse_lazy('chimp:show_mailing_list')

	def delete(self, request, *args, **kwargs):
		"""
		Calls the delete() method on the fetched object and then
		redirects to the success URL.
		"""
		self.object = self.get_object()
		success_url = self.get_success_url()
		self.object.delete()
		messages.success(request, 'Mailing List Deleted Successfully')
		return HttpResponseRedirect(success_url)



class EditUserProfileView(generic.FormView):		
	template_name='chimp/edit_user_profile.html'
	form_class=EditUserProfileForm
	success_url=reverse_lazy('chimp:dashboard')

	def form_valid(self,form):
		username = form.cleaned_data.get('username')
		name =form.cleaned_data.get('name')
		password=form.cleaned_data.get('password')
		business_name =form.cleaned_data.get('business_name')
		first_name=name.split()[0]
		last_name = ''
	
		try:
			last_name=name.split()[1]
		except:
			pass
	
		u = UserProfile.objects.get(user=self.request.user)
		u.business_name=business_name
		u.password=password
		u.business_name=business_name
		u.save()

		self.request.user.first_name = first_name
		self.request.user.last_name = last_name
		self.request.user.save()
		messages.success(self.request,'User Profile Edited successfully')
		return redirect(self.success_url)

